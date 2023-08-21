#!/usr/bin/env python39
import openpyxl as xl
from openpyxl import load_workbook
import requests
from math import log
import itertools
import re

#kernprof -lv CrossoutAPI.py

alpha = "abcdefghijklmnopqrstuvwxyz"
products = [list(itertools.product(''.join(alpha),repeat=r)) for r in range(1,4)]
products = products[0]+products[1]+products[2]
products = [''.join(list(i)) for i in products]

# 'Aa' -> 27; 1-based index; works with multiple digits & is case insensitive
# Premature optimization, you love to see it
# JAJAJA I DN"T EVEN USE THID FUNCTION
a_to_i = lambda a:sum([(int(j,36)-9)*26**(len(a)-i-1) for i,j in reversed(list(enumerate(a)))])
a_to_i_map = { i:a_to_i(i) for i in products }
a_to_i = lambda x:a_to_i_map[x.lower()]
del(products)

# 27 -> "AA"
i_to_a = lambda x: i_to_a((x-1)//26) + alpha[x%26-1] if x >= 27 else alpha[x%26-1]
### OPTIMISE THSI ^

# (1,6) -> "A6"; idk if works
xy_to_c = lambda x,y: i_to_a(x)+str(y)


# "A1:B3" -> [ (1,1), (2,1), (1,2), (2,2), (1,3), (2,3) ]
#@profile
def lrange(x):
    row_s,col_s,row_e,col_e = 0,0,0,0
    indx = x.index(':')
    l1 = sum([1 for i in x[:indx] if i.lower() in alpha]) # len of first letters
    l2 = sum([1 for i in x[indx+1:] if i.lower() not in alpha]) # len of second nums
    a,b,c,d = x[:l1], x[l1:indx], x[indx+1:-l2], x[-l2:] # the four parts ~[A,1,B,3]
    #print(a,b,c,d,list(range(int(a,36)-10,int(c,36)-9)))
    return [(j+1,i) for i in range(int(b),int(d)+1) for j in range(int(a,36)-10,int(c,36)-9)]

# Get updated data
def update_items(url="https://crossoutdb.com/api/v1/items"):
    response = requests.get(url)
    response_json = response.json()
    #print(response_json)
    with open("Items.txt","w+") as f:
        f.write(str(response_json))


# '=2*AC15 + AC17' -> '=MIN(X15:AC15)+MIN(X17:AC17)'; ChatGPT
#Oid_replace = lambda x: re.sub(r'AC(\d+)', lambda match: f"MIN(X{match.group(1)}:AC{match.group(1)})", x)
#Oid_replace = lambda x: re.sub(r'AC(\d+)', lambda match: f"W{match.group(1)}+MIN(X{match.group(1)}:AC{match.group(1)})", x)
#Oid_replace = lambda x: re.sub(r'AC(\d+)', lambda match: f"({f'W{match.group(1)}+MIN(X{match.group(1)}:AC{match.group(1)})'})", x)
Oid_replace = lambda x: re.sub(r'AB(\d+)', lambda match: f"AK{match.group(1)}", x)


file="Crossout.xlsx"#"D:\\Notes\\Python\\CrossoutAPI\\Crossout.xlsx"
# Get formulas
wb = xl.load_workbook(filename=file,data_only=False)
sheet = wb["Crafting+Market"]
Oid_range = lrange("T34:T194")
Oid_formulas = [sheet[xy_to_c(x,y)].value for x,y in Oid_range]
#print(Oid_formulas[:10])
#print([Oid_replace(i) for i in Oid_formulas[:10] if i is not None and i.startswith('=')])
for i in wb: print(i.title)
# Get values
wb = xl.load_workbook(filename=file,data_only=False)
sheet = wb["Crafting+Market"]

#print(sheet["y39"].value)

#@profile
def main():
    if 1:
        update_items()
    
    with open("Items.txt","r") as f:
        line = f.read()
    big_arr = eval(line)
    IDs = [i["id"] for i in big_arr]
    names = [i["name"] for i in big_arr]
    availnames = [i["availableName"] for i in big_arr]
    id_indx_map = { j['id']:i for i,j in enumerate(big_arr) } # works
    name_id_map = { i['name']:i['id'] for i in big_arr }
    avail_id_map = { i['availableName']:i['id'] for i in big_arr }
    
    if 0: # test
        for i in big_arr:
            if i["name"] == "LM-54 Chord":
                print('id:',i['id'])
        print(big_arr[1]["id"])

    if 1: # calculate Oid function
        for x,y in Oid_range:
            #cellval#print(x,y,name,sheet[eval(f"f'{i[1]}'")],item["sellPrice"])
            In = sheet[xy_to_c(x,y)].value
            #print(x,y,In)
            if In is not None and In.startswith('='):
                Out = Oid_replace(In)
                sheet[f"U{y}"] = Out
                #print(f"\tValid:V{y}",In,Out)
            #sheet[eval(f"f'{i[2]}'")] = item["buyPrice"]/100

    if 1: # write prices
        # [ [Input name, SO, BO, buyOffers] ]
        #           Store prices                         Material prices
        ranges = [ ["L8:L194","AA{y}","AB{y}","AI{y}"], ["C7:C22","D{y}","E{y}","A1"]]
        for i in ranges:
            for x,y in lrange(i[0]):
                name = sheet.cell(row=y, column=x).value
                #print((x,y), name, name in names)
                if name not in [None,"Name"]:
                    Map = name_id_map if name in names else (avail_id_map if name in availnames else None)
                    if Map != None:
                        item = big_arr[id_indx_map[Map[name]]]
                        if item["buyOrders"] >= 1:
                            #print(x,y,name,sheet[eval(f"f'{i[1]}'")],item["sellPrice"])
                            sheet[eval(f"f'{i[1]}'")] = item["sellPrice"]/100
                            sheet[eval(f"f'{i[2]}'")] = item["buyPrice"]/100
                            sheet[eval(f"f'{i[3]}'")] = item["buyOrders"]
                    else:
                        print("Doesn't work:",(x,y),name)
    
    
    wb.save(filename="outfile.xlsx")
    
    # print item keys
    #for i in big_arr[0].keys(): print(i)
main()







        
